"""
build_survey_template.py
---------------------------
Generates servqual_survey_template.xlsx:
  - "Instructions" tab explaining the field-collection process.
  - "Response Form" tab: one printable/fillable form (1 respondent) with all
    20 SERVQUAL items (Expectation + Perception, 1-7 Likert), pre-filled with
    one example row so a field surveyor knows the expected format.
  - "Raw Responses" tab: a long-format table (1 row per respondent) matching
    the exact column layout servqual_survey_analysis.py expects
    ("Dimension|item_index|E" / "...|P"), with the same example row and
    yellow-highlighted cells marking where new respondent rows should go.
"""
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import sys
sys.path.insert(0, ".")
try:
    from servqual_survey_analysis import SERVQUAL_ITEMS
except ImportError:
    from importlib import import_module

    SERVQUAL_ITEMS = import_module("servqual_survey_analysis (1)").SERVQUAL_ITEMS

NAVY = "1F3864"
YELLOW = "FFFF00"
LIGHT = "F2F2F2"

wb = openpyxl.Workbook()

# ---------------------------------------------------------------- Instructions
ws = wb.active
ws.title = "Instructions"
ws.column_dimensions["A"].width = 100
title = ws["A1"]
title.value = "Mini SERVQUAL Survey — Field Instructions"
title.font = Font(name="Arial", size=14, bold=True, color=NAVY)

instructions = [
    "",
    "Purpose: Triangulate the NLP/HDBSCAN cluster findings against direct customer feedback "
    "collected in person, addressing the evaluator's comment to visit outlets and confirm findings.",
    "",
    "Sample size: 10-15 respondents, at 1-2 outlets (choose outlets from chains with the highest "
    "review volume in the dashboard for maximum comparability).",
    "",
    "How to administer:",
    "  1. Approach customers as they finish their meal (avoid mid-meal interruptions).",
    "  2. Explain: 'This is a 2-minute academic survey on restaurant service quality, for a BITS "
    "Pilani MBA dissertation. Your responses are anonymous.'",
    "  3. For each of the 20 statements, ask for TWO ratings on a 1-7 scale (1 = Strongly Disagree, "
    "7 = Strongly Agree):",
    "       (a) EXPECTATION - 'Before you visit a restaurant like this, how much do you expect this "
    "to be true?'",
    "       (b) PERCEPTION - 'Thinking about today's visit, how true was this?'",
    "  4. Record both ratings in the 'Raw Responses' tab, one row per respondent, using the example "
    "row (highlighted) as a template — do not change the column headers.",
    "",
    "Analysis: once 10-15 rows are filled in, export/copy the 'Raw Responses' tab to CSV and load it "
    "into servqual_survey_analysis.py in place of generate_sample_responses().",
    "",
    "Dimensions covered: Tangibles, Reliability, Responsiveness, Assurance, Empathy (the 5 standard "
    "SERVQUAL dimensions; see the 'Response Form' tab for the exact statements).",
]
for i, line in enumerate(instructions, start=2):
    c = ws.cell(row=i, column=1, value=line)
    c.font = Font(name="Arial", size=11, bold=line.startswith(("Purpose", "Sample", "How", "Analysis", "Dimensions")))
    c.alignment = Alignment(wrap_text=True, vertical="top")

# ---------------------------------------------------------------- Response Form (printable, 1 respondent)
ws2 = wb.create_sheet("Response Form")
ws2.column_dimensions["A"].width = 4
ws2.column_dimensions["B"].width = 70
ws2.column_dimensions["C"].width = 14
ws2.column_dimensions["D"].width = 14

ws2["A1"] = "Mini SERVQUAL Survey — Response Form"
ws2["A1"].font = Font(name="Arial", size=13, bold=True, color=NAVY)
ws2.merge_cells("A1:D1")

ws2["A3"] = "Respondent #:"
ws2["B3"] = "[fill in]"
ws2["B3"].fill = PatternFill("solid", fgColor=YELLOW)
ws2["C3"] = "Outlet / Chain:"
ws2["D3"] = "[fill in]"
ws2["D3"].fill = PatternFill("solid", fgColor=YELLOW)

row = 5
ws2.cell(row=row, column=2, value="Statement (rate 1-7)").font = Font(bold=True)
ws2.cell(row=row, column=3, value="Expectation").font = Font(bold=True)
ws2.cell(row=row, column=4, value="Perception").font = Font(bold=True)
for col in range(2, 5):
    ws2.cell(row=row, column=col).fill = PatternFill("solid", fgColor=NAVY)
    ws2.cell(row=row, column=col).font = Font(bold=True, color="FFFFFF")
row += 1

for dim, items in SERVQUAL_ITEMS.items():
    ws2.cell(row=row, column=1, value=dim).font = Font(bold=True, italic=True, color=NAVY)
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws2.cell(row=row, column=1).fill = PatternFill("solid", fgColor=LIGHT)
    row += 1
    for item in items:
        ws2.cell(row=row, column=2, value=item).alignment = Alignment(wrap_text=True)
        ws2.cell(row=row, column=3, value="[1-7]").fill = PatternFill("solid", fgColor=YELLOW)
        ws2.cell(row=row, column=4, value="[1-7]").fill = PatternFill("solid", fgColor=YELLOW)
        row += 1

ws2.freeze_panes = "A6"
ws2.page_setup.fitToWidth = 1
ws2.page_setup.fitToHeight = 0
ws2.sheet_properties.pageSetUpPr.fitToPage = True
ws2.page_setup.orientation = "landscape"

# ---------------------------------------------------------------- Raw Responses (analysis-ready, long format)
ws3 = wb.create_sheet("Raw Responses")
headers = ["respondent_id", "outlet"]
for dim, items in SERVQUAL_ITEMS.items():
    for i in range(len(items)):
        headers.append(f"{dim}|{i}|E")
        headers.append(f"{dim}|{i}|P")

for c_idx, h in enumerate(headers, start=1):
    cell = ws3.cell(row=1, column=c_idx, value=h)
    cell.font = Font(bold=True, color="FFFFFF", size=9)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    ws3.column_dimensions[get_column_letter(c_idx)].width = 16

# one example row so surveyors see the expected format
example = ["R1", "A2B - Anna Nagar"]
for dim, items in SERVQUAL_ITEMS.items():
    for i in range(len(items)):
        example += [6, 5]  # example Likert values
for c_idx, val in enumerate(example, start=1):
    cell = ws3.cell(row=2, column=c_idx, value=val)
    cell.fill = PatternFill("solid", fgColor=YELLOW)

# 14 more blank rows ready for the remaining respondents (target n=15)
for r in range(3, 17):
    ws3.cell(row=r, column=1, value=f"R{r-1}")

ws3.freeze_panes = "A2"

wb.save("servqual_survey_template.xlsx")
print("Saved -> servqual_survey_template.xlsx")
