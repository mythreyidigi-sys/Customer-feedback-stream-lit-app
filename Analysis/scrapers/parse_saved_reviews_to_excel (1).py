"""
Parses REVIEW pages you've saved locally from Zomato and Tripadvisor into an
Excel sheet with columns: Restaurant, Branch, Reviews, Source.

WHY THIS APPROACH:
Zomato's robots.txt disallows automated fetching of its review pages, and
both Zomato's and Tripadvisor's Terms of Service prohibit scraping. A script
that repeatedly requests those pages over the network would be circumventing
that restriction. This script instead works on HTML *you* already saved from
your own browser session:

  1. Open the review page in your browser (log in if needed).
  2. Scroll down / click "Load more" until all the reviews you want are
     visible on the page.
  3. Right-click -> "Save As" -> choose "Webpage, HTML only" (or Ctrl+S).
  4. Save it into the same folder as this script (or update the paths below).
  5. Run this script.

HOW TO GET THE PATHS RIGHT:
Update ZOMATO_HTML_PATH and TRIPADVISOR_HTML_PATH below to point at your
saved files.

NOTE ON RELIABILITY:
Zomato and Tripadvisor frequently change their page structure and class
names, and much of Zomato's content loads via JavaScript, so a saved
"HTML only" file may not contain the review text if it wasn't rendered at
save time. If a selector below returns 0 reviews, open the saved file in a
text editor, search for a snippet of review text you can see on-screen, and
adjust the CSS selectors marked "ADJUST THIS" to match what's actually in
the file.
"""

import re
from pathlib import Path

from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# --- CONFIGURE THESE PATHS ---
ZOMATO_HTML_PATH = "zomato_geetham_tnagar_reviews.html"
TRIPADVISOR_HTML_PATH = "tripadvisor_geetham_reviews.html"
OUTPUT_XLSX_PATH = "/mnt/user-data/outputs/geetham_reviews_from_saved_pages.xlsx"

RESTAURANT_NAME = "Geetham Veg Restaurant"


def clean_text(text: str) -> str:
    """Collapse whitespace/newlines."""
    return re.sub(r"\s+", " ", text or "").strip()


def parse_zomato(html_path: str, branch: str):
    """
    Extract review text from a saved Zomato reviews page.
    ADJUST THIS: Zomato's review text typically sits in <p> tags inside a
    review card container. Inspect the saved HTML (Ctrl+F for a phrase from
    a review you can see on screen) and update the selector if this misses.
    """
    reviews = []
    path = Path(html_path)
    if not path.exists():
        print(f"[Zomato] File not found: {html_path} -- skipping.")
        return reviews

    soup = BeautifulSoup(path.read_text(encoding="utf-8", errors="ignore"), "html.parser")

    # Common Zomato review-text containers (subject to change -- adjust as needed)
    candidates = soup.select("p.res-review-text, div.rev-text, p[class*='review']")

    seen = set()
    for tag in candidates:
        text = clean_text(tag.get_text())
        if len(text) > 20 and text not in seen:
            seen.add(text)
            reviews.append((RESTAURANT_NAME, branch, text, "Zomato"))

    if not reviews:
        print("[Zomato] No reviews matched the selectors -- open the saved "
              "HTML and update the CSS selectors in parse_zomato().")
    return reviews


def parse_tripadvisor(html_path: str, branch: str):
    """
    Extract review text from a saved Tripadvisor reviews page.
    ADJUST THIS: Tripadvisor typically renders each review's text in a <q>
    or <span> tag inside a review card (data-test-target="review-text" or
    similar). Update the selector if this misses.
    """
    reviews = []
    path = Path(html_path)
    if not path.exists():
        print(f"[Tripadvisor] File not found: {html_path} -- skipping.")
        return reviews

    soup = BeautifulSoup(path.read_text(encoding="utf-8", errors="ignore"), "html.parser")

    candidates = soup.select(
        "[data-test-target='review-text'], q.QewHA, span.yCeTE, div.biGQs.orRIx"
    )

    seen = set()
    for tag in candidates:
        text = clean_text(tag.get_text())
        if len(text) > 20 and text not in seen:
            seen.add(text)
            reviews.append((RESTAURANT_NAME, branch, text, "Tripadvisor"))

    if not reviews:
        print("[Tripadvisor] No reviews matched the selectors -- open the "
              "saved HTML and update the CSS selectors in parse_tripadvisor().")
    return reviews


def build_excel(rows, output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reviews"

    headers = ["Restaurant", "Branch", "Reviews", "Source"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="Arial", bold=True, color="FFFFFF")

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, (restaurant, branch, review, source) in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=restaurant).font = Font(name="Arial")
        ws.cell(row=row_idx, column=2, value=branch).font = Font(name="Arial")
        review_cell = ws.cell(row=row_idx, column=3, value=review)
        review_cell.font = Font(name="Arial")
        review_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row_idx, column=4, value=source).font = Font(name="Arial")

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 90
    ws.column_dimensions["D"].width = 14
    ws.freeze_panes = "A2"

    wb.save(output_path)
    print(f"Saved {len(rows)} reviews to {output_path}")


if __name__ == "__main__":
    all_rows = []
    all_rows += parse_zomato(ZOMATO_HTML_PATH, branch="T. Nagar")
    all_rows += parse_tripadvisor(TRIPADVISOR_HTML_PATH, branch="Medavakkam")

    if not all_rows:
        print("\nNo reviews were extracted. Make sure you've saved the pages "
              "locally and updated ZOMATO_HTML_PATH / TRIPADVISOR_HTML_PATH "
              "above, then re-run.")
    else:
        build_excel(all_rows, OUTPUT_XLSX_PATH)
